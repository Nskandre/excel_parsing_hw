from collections import Counter
from openpyxl import load_workbook
import pandas


def make_report(log_file, report_template, report_output):
    
    wb = load_workbook(report_template) # open existing workbook

    ws = wb.active # trans in variable active sheet

    read_logs = pandas.read_excel(log_file)
    logs_dict = read_logs.to_dict(orient='records')

    browsers = {} # словарь всех браузеров с посещением помесячно
    visits_of_browsers = {} # словарь всех браузеров с посещениями за весь период
    rating_of_browsers = {} # 7 самых популярных браузеров
    res_browsers = {} # 7 самых популярных браузеров с посещениями по месяцам

    for i in logs_dict:
        if i['Браузер'] in browsers:
            browsers[i['Браузер']][str(i['Дата посещения'].month)] = browsers[i['Браузер']].get(
                str(i['Дата посещения'].month), 0) + 1
        else:
            browsers[i['Браузер']] = {
                str(i['Дата посещения'].month): browsers.get(str(i['Дата посещения'].month), 0) + 1}

        visits_of_browsers[i['Браузер']] = visits_of_browsers.get(i['Браузер'], 0) + 1

    # сортировка посещений браузеров в обратном порядке
    sorted_visits_of_browser = sorted(visits_of_browsers.values(), reverse=True)

    for i in sorted_visits_of_browser:
        for browser in visits_of_browsers.keys():
            if visits_of_browsers[browser] == i:
                rating_of_browsers[browser] = i
            if len(rating_of_browsers) >= 7:
                break
        if len(rating_of_browsers) >= 7:
            break

    for browser in rating_of_browsers.keys():
        res_browsers[browser] = browsers[browser]

    goods = {} # словарь всех товаров с их покупками помесячно
    buyings_of_goods = {} # словарь всех товаров с их покупками за весь период
    rating_of_goods = {} # 7 самых популярных товаров
    res_goods = {} # 7 самых популярных товаров по месяцам

    for i in logs_dict:
        for j in i['Купленные товары'].split(','):
            if j.rstrip() in goods:
                goods[j.rstrip()][str(i['Дата посещения'].month)] = goods[j.rstrip()].get(
                    str(i['Дата посещения'].month), 0) + 1
            else:
                goods[j.rstrip()] = {str(i['Дата посещения'].month): goods.get(str(i['Дата посещения'].month), 0) + 1}
            buyings_of_goods[j.rstrip()] = buyings_of_goods.get(j.rstrip(), 0) + 1

    # сортировка товаров по покупкам в обратном порядке
    sorted_buyings_of_goods = sorted(buyings_of_goods.values(), reverse=True)


    for i in sorted_buyings_of_goods:
        for good in buyings_of_goods.keys():
            if buyings_of_goods[good] == i:
                rating_of_goods[good] = i
            if len(rating_of_goods) >= 7:
                break
        if len(rating_of_goods) >= 7:
            break

    for good in rating_of_goods.keys():
        res_goods[good] = goods[good]

    mens_choice = [] # список предпочтений мужчин
    for i in logs_dict:
        if i['Пол'] == 'м':
            for good in i['Купленные товары'].split(','):
                mens_choice.append(good.rstrip())

    mens_goods_dict = Counter(mens_choice) # словарь с количеством покупок предпочтений мужчин

    women_choice = [] # список предпочтений женщин
    for i in logs_dict:
        if i['Пол'] == 'ж':
            for good in i['Купленные товары'].split(','):
                women_choice.append(good.rstrip())

    women_goods_dict = Counter(women_choice) # словарь с количеством покупок предпочтений женщин

    mens_popular_goods = ''
    mens_useless_goods = ''
    for good in mens_goods_dict:
        if len(mens_popular_goods) == 0 or mens_goods_dict[good] > mens_goods_dict[mens_popular_goods]:
            mens_popular_goods = good
        elif len(mens_useless_goods) == 0 or mens_goods_dict[good] < mens_goods_dict[mens_useless_goods]:
            mens_useless_goods = good

    women_popular_goods = ''
    women_useless_goods = ''
    for good in women_goods_dict:
        if len(women_popular_goods) == 0 or women_goods_dict[good] > women_goods_dict[women_popular_goods]:
            women_popular_goods = good
        elif len(women_useless_goods) == 0 or women_goods_dict[good] < women_goods_dict[women_useless_goods]:
            women_useless_goods = good


    # заполнение эксел файла - топ 7 браузеров по убыванию с помесячными посещениями
    row_browsers = 5
    for key in res_browsers:
        _ = ws.cell(column=1, row=row_browsers, value=key)
        for value in res_browsers[key]:
            _ = ws.cell(column=int(value)+1, row=row_browsers, value=res_browsers[key][value])
        row_browsers += 1

    # заполнение эксел файла - топ 7 браузеров по убыванию с помесячными посещениями
    row_goods = 19
    for key in res_goods:
        _ = ws.cell(column=1, row=row_goods, value=key)
        for value in res_goods[key]:
            _ = ws.cell(column=int(value)+1, row=row_goods, value=res_goods[key][value])
        row_goods += 1

    # популярные и непопулярные товары у мужчин и женщин
    ws['B31'] = mens_popular_goods
    ws['B32'] = women_popular_goods
    ws['B33'] = mens_useless_goods
    ws['B34'] = women_useless_goods


    wb.save(report_output)


make_report('logs.xlsx', 'report_template.xlsx', 'report.xlsx')